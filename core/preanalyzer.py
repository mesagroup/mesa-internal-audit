"""
Pre-analisi deterministica per controlli strutturati.

Per C01 (PR/PO Authorization & SoD) e C03 (Payment Terms Alignment),
i check point numerici vengono calcolati in Python e iniettati nel prompt
come fatti verificati. L'LLM riceve i risultati già computati e deve solo
generare il testo narrativo.
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Tuple


def preanalyze_c01(file_paths: List[Path]) -> str:
    """
    Cerca un file Excel con foglio PR_PO_Register tra i documenti caricati.
    Se trovato, calcola deterministicamente ogni check point per ogni riga
    e restituisce un blocco di testo da iniettare nel prompt.
    """
    for path in file_paths:
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        result = _analyze_excel(path)
        if result:
            return result
    return ""


def _analyze_excel(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""

    wb = load_workbook(str(path), read_only=True, data_only=True)
    if "PR_PO_Register" not in wb.sheetnames:
        return ""

    ws_pr = wb["PR_PO_Register"]
    rows = list(ws_pr.iter_rows(values_only=True))
    if not rows:
        return ""

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]

    def col(row, name):
        try:
            return row[headers.index(name)]
        except ValueError:
            return None

    # Leggi DoA se presente
    doa_map: dict[str, float] = {}
    if "Delegation_of_Authority" in wb.sheetnames:
        ws_doa = wb["Delegation_of_Authority"]
        doa_rows = list(ws_doa.iter_rows(values_only=True))
        for r in doa_rows[1:]:
            if r and r[0] and r[1]:
                doa_map[str(r[0]).strip()] = float(r[1])

    lines = ["## PRE-ANALISI DETERMINISTICA (calcolata in Python — usa questi risultati come fatti certi)\n"]
    lines.append(
        "IMPORTANTE: I risultati seguenti sono stati calcolati matematicamente da codice Python, "
        "NON devi ricalcolarli. Usali direttamente per determinare lo status di ogni check point.\n"
    )

    cp1_issues: List[str] = []
    cp2_issues: List[str] = []
    cp3_issues: List[str] = []
    cp4_issues: List[str] = []

    for row in data_rows:
        if not any(v is not None for v in row):
            continue

        pr_num = col(row, "PR_Number") or "?"
        amount = _to_float(col(row, "PR_Amount_EUR"))
        budget = _to_float(col(row, "Budget_Available_EUR"))
        doa_thresh = _to_float(col(row, "DoA_Threshold_EUR"))
        pr_date = col(row, "PR_Date")
        po_date = col(row, "PO_Date")
        pr_creator = str(col(row, "PR_Creator") or "").strip()
        po_creator = str(col(row, "PO_Creator") or "").strip()
        approver = str(col(row, "PR_Approver") or "").strip()

        # CP1: importo <= soglia DoA
        if amount is not None and doa_thresh is not None:
            ok = amount <= doa_thresh
            tag = "CONFORME" if ok else "NON CONFORME"
            detail = f"{pr_num}: importo {amount:,.0f} {'<=' if ok else '>'} soglia DoA {doa_thresh:,.0f} → {tag} (approvatore: {approver})"
            lines.append(f"  [CP1] {detail}")
            if not ok:
                cp1_issues.append(detail)

        # CP2: PO_Date > PR_Date
        if pr_date is not None and po_date is not None:
            ok = po_date > pr_date
            tag = "CONFORME" if ok else "NON CONFORME"
            lines.append(f"  [CP2] {pr_num}: PR_Date={_fmt_date(pr_date)}, PO_Date={_fmt_date(po_date)} → {tag}")
            if not ok:
                cp2_issues.append(f"{pr_num}: PO emesso PRIMA della PR")

        # CP3: importo <= budget
        if amount is not None and budget is not None:
            ok = amount <= budget
            tag = "CONFORME" if ok else "NON CONFORME"
            detail = f"{pr_num}: importo {amount:,.0f} {'<=' if ok else '>'} budget {budget:,.0f} → {tag}"
            lines.append(f"  [CP3] {detail}")
            if not ok:
                cp3_issues.append(detail)

        # CP4: PR_Creator != PO_Creator
        if pr_creator and po_creator:
            ok = pr_creator.lower() != po_creator.lower()
            tag = "CONFORME" if ok else "NON CONFORME"
            lines.append(f"  [CP4] {pr_num}: PR_Creator='{pr_creator}' vs PO_Creator='{po_creator}' → {tag}")
            if not ok:
                cp4_issues.append(f"{pr_num}: stesso soggetto '{pr_creator}'")

    lines.append("")
    lines.append("### SINTESI PRE-ANALISI:")
    lines.append(f"  CP1 (DoA): {'TUTTI CONFORMI' if not cp1_issues else 'NON CONFORME — ' + '; '.join(cp1_issues)}")
    lines.append(f"  CP2 (date): {'TUTTI CONFORMI' if not cp2_issues else 'NON CONFORME — ' + '; '.join(cp2_issues)}")
    lines.append(f"  CP3 (budget): {'TUTTI CONFORMI' if not cp3_issues else 'NON CONFORME — ' + '; '.join(cp3_issues)}")
    lines.append(f"  CP4 (SoD): {'TUTTI CONFORMI' if not cp4_issues else 'NON CONFORME — ' + '; '.join(cp4_issues)}")
    lines.append("")
    lines.append(
        "Usa ESCLUSIVAMENTE questi risultati per determinare lo status JSON. "
        "Non ricalcolare, non reinterpretare. Scrivi solo il testo narrativo (evidence, issue, mitigation)."
    )

    return "\n".join(lines)


def preanalyze_c03(file_paths: List[Path]) -> str:
    """
    Cerca un file Excel con foglio Invoices_Q1_2026 e Contract_Payment_Terms.
    Calcola deterministicamente i 3 check point di C03.
    """
    for path in file_paths:
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        result = _analyze_excel_c03(path)
        if result:
            return result
    return ""


def _analyze_excel_c03(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""

    wb = load_workbook(str(path), read_only=True, data_only=True)
    if "Invoices_Q1_2026" not in wb.sheetnames:
        return ""

    # Leggi contratti
    contract_map: dict[str, str] = {}
    if "Contract_Payment_Terms" in wb.sheetnames:
        ws_c = wb["Contract_Payment_Terms"]
        rows_c = list(ws_c.iter_rows(values_only=True))
        for r in rows_c[1:]:
            if r and r[0] and r[2]:
                contract_map[str(r[0]).strip()] = str(r[2]).strip()

    ws_inv = wb["Invoices_Q1_2026"]
    rows = list(ws_inv.iter_rows(values_only=True))
    if not rows:
        return ""
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    def col(row, name):
        try:
            return row[headers.index(name)]
        except ValueError:
            return None

    lines = ["## PRE-ANALISI DETERMINISTICA C03 (calcolata in Python — usa questi risultati come fatti certi)\n"]
    lines.append(
        "IMPORTANTE: I risultati seguenti sono stati calcolati matematicamente. "
        "Non ricalcolarli. Usali per determinare lo status di ogni check point.\n"
    )

    cp1_issues: List[str] = []
    cp2_missing: List[str] = []
    cp3_issues: List[str] = []

    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue

        supplier = str(col(row, "Supplier_Name") or "").strip()
        inv_num = str(col(row, "Invoice_Number") or "?").strip()
        due_date = col(row, "Due_Date_Calculated")
        pay_date = col(row, "Actual_Payment_Date")
        days_late = col(row, "Days_Late")

        # CP1: pagato entro la scadenza (Days_Late <= 0)
        if days_late is not None:
            try:
                dl = float(days_late)
                ok = dl <= 0
                tag = "CONFORME" if ok else f"NON CONFORME ({int(dl):+d} giorni)"
                lines.append(
                    f"  [CP1] {inv_num} ({supplier}): scadenza={_fmt_date(due_date)}, "
                    f"pagato={_fmt_date(pay_date)}, Days_Late={int(dl):+d} → {tag}"
                )
                if not ok:
                    cp1_issues.append(f"{inv_num}: {int(dl):+d} giorni oltre scadenza")
            except (TypeError, ValueError):
                pass

        # CP2: contratto con PT definiti
        if supplier:
            has_contract = supplier in contract_map
            tag = "CONFORME" if has_contract else "NON CONFORME"
            pt = contract_map.get(supplier, "non trovato")
            lines.append(f"  [CP2] {supplier}: PT contrattuale='{pt}' → {tag}")
            if not has_contract:
                cp2_missing.append(supplier)

        # CP3: discrepanze documentate — applicabile solo se Days_Late > 0
        if days_late is not None:
            try:
                dl = float(days_late)
                if dl > 0:
                    cp3_issues.append(f"{inv_num}: ritardo {int(dl):+d}gg, verifica documentazione")
            except (TypeError, ValueError):
                pass

    lines.append("")
    lines.append("### SINTESI PRE-ANALISI:")
    lines.append(f"  CP1 (rispetto scadenze): {'TUTTI CONFORMI' if not cp1_issues else 'NON CONFORME — ' + '; '.join(cp1_issues)}")
    lines.append(f"  CP2 (PT da contratto/PO): {'TUTTI CONFORMI' if not cp2_missing else 'NON CONFORME — PT mancanti per: ' + '; '.join(set(cp2_missing))}")
    if cp3_issues:
        lines.append(f"  CP3 (discrepanze documentate): VERIFICARE — {'; '.join(cp3_issues)}")
    else:
        lines.append("  CP3 (discrepanze documentate): NESSUNA DISCREPANZA RILEVATA — check point non applicabile (tutti i pagamenti sono puntuali o anticipati)")
    lines.append("")
    lines.append(
        "Usa ESCLUSIVAMENTE questi risultati per lo status JSON. "
        "Non ricalcolare, non reinterpretare. Scrivi solo il testo narrativo."
    )

    return "\n".join(lines)


def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _fmt_date(val) -> str:
    if val is None:
        return "N/D"
    try:
        return val.strftime("%Y-%m-%d")
    except AttributeError:
        return str(val)
