"""
Generazione del report di audit finale in formato PDF ed Excel.
Usa reportlab per il PDF e openpyxl per l'Excel.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

from .llm_verifier import VerificationResult


STATUS_COLORS = {
    "conforme": colors.HexColor("#1f7a1f"),
    "non_conforme": colors.HexColor("#b22222"),
    "parziale": colors.HexColor("#c78a00"),
    "non_verificabile": colors.HexColor("#666666"),
}


def _styles():
    ss = getSampleStyleSheet()
    ss.add(
        ParagraphStyle(
            name="H1Custom",
            fontName="Helvetica-Bold",
            fontSize=18,
            spaceAfter=12,
            textColor=colors.HexColor("#222222"),
        )
    )
    ss.add(
        ParagraphStyle(
            name="H2Custom",
            fontName="Helvetica-Bold",
            fontSize=14,
            spaceBefore=14,
            spaceAfter=8,
            textColor=colors.HexColor("#333333"),
        )
    )
    ss.add(
        ParagraphStyle(
            name="H3Custom",
            fontName="Helvetica-Bold",
            fontSize=11,
            spaceBefore=8,
            spaceAfter=4,
            textColor=colors.HexColor("#444444"),
        )
    )
    ss.add(
        ParagraphStyle(
            name="BodyCustom",
            fontName="Helvetica",
            fontSize=10,
            leading=13,
            spaceAfter=6,
        )
    )
    ss.add(
        ParagraphStyle(
            name="SmallCustom",
            fontName="Helvetica",
            fontSize=9,
            leading=11,
            textColor=colors.HexColor("#555555"),
        )
    )
    return ss


def _status_badge(status: str) -> Paragraph:
    color = STATUS_COLORS.get(status, colors.black)
    label = status.replace("_", " ").upper()
    return Paragraph(
        f'<font color="{color.hexval()}"><b>{label}</b></font>',
        getSampleStyleSheet()["Normal"],
    )


def _executive_summary_table(results: List[VerificationResult], styles) -> Table:
    data = [["ID", "Controllo", "Esito", "Check point NC"]]
    for r in results:
        nc = sum(1 for cp in r.check_points if cp.status == "non_conforme")
        total = len(r.check_points)
        data.append(
            [
                r.control_id,
                Paragraph(r.control_title, styles["BodyCustom"]),
                _status_badge(r.overall_status),
                f"{nc}/{total}",
            ]
        )
    tbl = Table(data, colWidths=[1.5 * cm, 8.5 * cm, 3.5 * cm, 3 * cm])
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (2, 0), (3, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return tbl


def _check_points_table(result: VerificationResult, styles) -> Table:
    data = [["#", "Check Point", "Esito", "Evidenza / Issue"]]
    for i, cp in enumerate(result.check_points, start=1):
        issue_text = cp.evidence
        if cp.issue:
            issue_text += f"<br/><i>Issue:</i> {cp.issue}"
        data.append(
            [
                str(i),
                Paragraph(cp.check_point, styles["BodyCustom"]),
                _status_badge(cp.status),
                Paragraph(issue_text, styles["SmallCustom"]),
            ]
        )
    tbl = Table(data, colWidths=[0.8 * cm, 6.5 * cm, 3 * cm, 6.2 * cm], repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (2, 0), (2, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9f9f9")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return tbl


def generate_pdf_report(
    results: List[VerificationResult],
    output_path: str | Path,
    title: str = "Internal Audit Report — Procure-to-Pay",
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=1.8 * cm,
        leftMargin=1.8 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
    )
    styles = _styles()
    story = []

    # --- Header ---
    story.append(Paragraph(title, styles["H1Custom"]))
    story.append(
        Paragraph(
            f"Data esecuzione: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles["SmallCustom"],
        )
    )
    story.append(Spacer(1, 0.4 * cm))

    # --- Executive summary ---
    story.append(Paragraph("Executive Summary", styles["H2Custom"]))

    total = len(results)
    conformi = sum(1 for r in results if r.overall_status == "conforme")
    non_conformi = sum(1 for r in results if r.overall_status == "non_conforme")
    parziali = total - conformi - non_conformi

    story.append(
        Paragraph(
            f"Sono stati eseguiti <b>{total}</b> controlli: "
            f"<font color='#1f7a1f'><b>{conformi}</b> conformi</font>, "
            f"<font color='#b22222'><b>{non_conformi}</b> non conformi</font>, "
            f"<font color='#c78a00'><b>{parziali}</b> parziali</font>.",
            styles["BodyCustom"],
        )
    )
    story.append(Spacer(1, 0.3 * cm))
    story.append(_executive_summary_table(results, styles))

    # --- Dettaglio per ogni controllo ---
    for r in results:
        story.append(PageBreak())
        story.append(
            Paragraph(f"Controllo {r.control_id} — {r.control_title}", styles["H2Custom"])
        )
        story.append(
            Paragraph(
                f"<b>Esito complessivo:</b> "
                f"<font color='{STATUS_COLORS.get(r.overall_status, colors.black).hexval()}'>"
                f"<b>{r.overall_status.replace('_', ' ').upper()}</b></font>",
                styles["BodyCustom"],
            )
        )
        story.append(Spacer(1, 0.2 * cm))

        story.append(Paragraph("Sintesi", styles["H3Custom"]))
        story.append(Paragraph(r.summary or "—", styles["BodyCustom"]))

        story.append(Paragraph("Dettaglio Check Points", styles["H3Custom"]))
        story.append(_check_points_table(r, styles))

        if r.mitigation_plan:
            story.append(Spacer(1, 0.3 * cm))
            story.append(Paragraph("Piano di Mitigazione", styles["H3Custom"]))
            for i, m in enumerate(r.mitigation_plan, start=1):
                story.append(Paragraph(f"<b>{i}.</b> {m}", styles["BodyCustom"]))

    doc.build(story)
    return output_path


# ── Status label helpers ────────────────────────────────────────────────────

_STATUS_LABELS = {
    "conforme": "CONFORME",
    "non_conforme": "NON CONFORME",
    "parziale": "PARZIALE",
    "non_verificabile": "NON VERIFICABILE",
}

_STATUS_FILLS = {
    "conforme": PatternFill("solid", fgColor="1f7a1f"),
    "non_conforme": PatternFill("solid", fgColor="b22222"),
    "parziale": PatternFill("solid", fgColor="c78a00"),
    "non_verificabile": PatternFill("solid", fgColor="666666"),
}

_HEADER_FILL = PatternFill("solid", fgColor="2c3e50")
_SUBHEADER_FILL = PatternFill("solid", fgColor="34495e")
_ALT_FILL = PatternFill("solid", fgColor="f5f5f5")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

_THIN_BORDER = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)


def _xl_header(cell, text: str):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _THIN_BORDER


def _xl_subheader(cell, text: str):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = _SUBHEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _THIN_BORDER


def _xl_cell(cell, text: str, alt: bool = False, bold: bool = False):
    cell.value = text
    cell.font = Font(bold=bold, size=10)
    cell.fill = _ALT_FILL if alt else _WHITE_FILL
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = _THIN_BORDER


def _xl_status_cell(cell, status: str, alt: bool = False):
    label = _STATUS_LABELS.get(status, status.upper())
    cell.value = label
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = _STATUS_FILLS.get(status, PatternFill("solid", fgColor="333333"))
    cell.alignment = Alignment(horizontal="center", vertical="top")
    cell.border = _THIN_BORDER


def generate_excel_report(
    results: List[VerificationResult],
    output_path: str | Path,
    title: str = "Internal Audit Report — Procure-to-Pay",
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Executive Summary ──────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"

    ws_summary.merge_cells("A1:F1")
    title_cell = ws_summary["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1a252f")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 30

    ws_summary.merge_cells("A2:F2")
    date_cell = ws_summary["A2"]
    date_cell.value = f"Data esecuzione: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    date_cell.font = Font(italic=True, size=10, color="555555")
    date_cell.alignment = Alignment(horizontal="center")
    ws_summary.row_dimensions[2].height = 18

    total = len(results)
    conformi = sum(1 for r in results if r.overall_status == "conforme")
    non_conformi = sum(1 for r in results if r.overall_status == "non_conforme")
    parziali = total - conformi - non_conformi

    ws_summary.merge_cells("A3:F3")
    stats_cell = ws_summary["A3"]
    stats_cell.value = (
        f"Controlli eseguiti: {total}  |  "
        f"Conformi: {conformi}  |  "
        f"Non conformi: {non_conformi}  |  "
        f"Parziali: {parziali}"
    )
    stats_cell.font = Font(bold=True, size=10)
    stats_cell.alignment = Alignment(horizontal="center")
    ws_summary.row_dimensions[3].height = 18

    headers = ["ID", "Controllo", "Esito", "CP Totali", "CP Non Conformi", "CP Conformi"]
    for col, h in enumerate(headers, start=1):
        _xl_header(ws_summary.cell(row=5, column=col), h)
    ws_summary.row_dimensions[5].height = 22

    for row_idx, r in enumerate(results, start=6):
        alt = (row_idx % 2 == 0)
        nc = sum(1 for cp in r.check_points if cp.status == "non_conforme")
        ok = sum(1 for cp in r.check_points if cp.status == "conforme")
        total_cp = len(r.check_points)
        _xl_cell(ws_summary.cell(row=row_idx, column=1), r.control_id, alt, bold=True)
        _xl_cell(ws_summary.cell(row=row_idx, column=2), r.control_title, alt)
        _xl_status_cell(ws_summary.cell(row=row_idx, column=3), r.overall_status, alt)
        _xl_cell(ws_summary.cell(row=row_idx, column=4), str(total_cp), alt)
        _xl_cell(ws_summary.cell(row=row_idx, column=5), str(nc), alt)
        _xl_cell(ws_summary.cell(row=row_idx, column=6), str(ok), alt)
        ws_summary.row_dimensions[row_idx].height = 20

    col_widths = [10, 50, 20, 12, 20, 15]
    for col, w in enumerate(col_widths, start=1):
        ws_summary.column_dimensions[get_column_letter(col)].width = w

    ws_summary.freeze_panes = "A6"

    # ── Sheet 2: Check Points dettaglio ────────────────────────────────────
    ws_cp = wb.create_sheet("Check Points")
    cp_headers = ["Controllo ID", "Titolo Controllo", "#", "Check Point", "Esito", "Evidenza", "Issue"]
    for col, h in enumerate(cp_headers, start=1):
        _xl_header(ws_cp.cell(row=1, column=col), h)
    ws_cp.row_dimensions[1].height = 22

    cp_row = 2
    for r in results:
        for i, cp in enumerate(r.check_points, start=1):
            alt = (cp_row % 2 == 0)
            _xl_cell(ws_cp.cell(row=cp_row, column=1), r.control_id, alt, bold=True)
            _xl_cell(ws_cp.cell(row=cp_row, column=2), r.control_title, alt)
            _xl_cell(ws_cp.cell(row=cp_row, column=3), str(i), alt)
            _xl_cell(ws_cp.cell(row=cp_row, column=4), cp.check_point, alt)
            _xl_status_cell(ws_cp.cell(row=cp_row, column=5), cp.status, alt)
            _xl_cell(ws_cp.cell(row=cp_row, column=6), cp.evidence or "", alt)
            _xl_cell(ws_cp.cell(row=cp_row, column=7), cp.issue or "", alt)
            ws_cp.row_dimensions[cp_row].height = 40
            cp_row += 1

    cp_col_widths = [14, 40, 5, 50, 20, 40, 40]
    for col, w in enumerate(cp_col_widths, start=1):
        ws_cp.column_dimensions[get_column_letter(col)].width = w

    ws_cp.freeze_panes = "A2"

    # ── Sheet 3: Piani di mitigazione ──────────────────────────────────────
    ws_mit = wb.create_sheet("Piani di Mitigazione")
    mit_headers = ["Controllo ID", "Titolo Controllo", "Esito", "#", "Azione di Mitigazione"]
    for col, h in enumerate(mit_headers, start=1):
        _xl_header(ws_mit.cell(row=1, column=col), h)
    ws_mit.row_dimensions[1].height = 22

    mit_row = 2
    for r in results:
        if not r.mitigation_plan:
            continue
        for i, action in enumerate(r.mitigation_plan, start=1):
            alt = (mit_row % 2 == 0)
            _xl_cell(ws_mit.cell(row=mit_row, column=1), r.control_id, alt, bold=True)
            _xl_cell(ws_mit.cell(row=mit_row, column=2), r.control_title, alt)
            _xl_status_cell(ws_mit.cell(row=mit_row, column=3), r.overall_status, alt)
            _xl_cell(ws_mit.cell(row=mit_row, column=4), str(i), alt)
            _xl_cell(ws_mit.cell(row=mit_row, column=5), action, alt)
            ws_mit.row_dimensions[mit_row].height = 40
            mit_row += 1

    mit_col_widths = [14, 40, 20, 5, 80]
    for col, w in enumerate(mit_col_widths, start=1):
        ws_mit.column_dimensions[get_column_letter(col)].width = w

    ws_mit.freeze_panes = "A2"

    wb.save(output_path)
    return output_path
